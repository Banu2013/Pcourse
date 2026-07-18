"""
=========================================================
BaavaTech BOM Automation Tool
Version : 1.0

Developed by:
BaavaTech

Author:
Banumathi

Copyright © 2026 BaavaTech.
All Rights Reserved.
=========================================================
"""
import os
import logging
from copy import copy

import openpyxl
from openpyxl import load_workbook

from config import (
    VSE_SHEET,
    DATA_START_ROW,
    LOG_FILE
)

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)


class VSEGenerator:
   
    def __init__(self, cleaned_file, template_file, generated_vse):

        self.cleaned_file = cleaned_file
        self.template_file = template_file
        self.generated_vse = generated_vse

        self.bom_data = None
        self.bom_sheet = None
        self.headers = []

        self.workbook = None
        self.sheet = None

        self.column_mapping = {

            "Lvl": "Lvl",
            "Item #": "Seq",
            "VSE P/N": "VSE Item",
            "Qty": "Qty",
            "Customer P/N": "Cust Item",
            "Rev": "Rev",
            "Description": "Description",
            "Reference Designator": "Reference Designator",
            "UOM": "UOM",
            "Manufacturer": "Manufacturer",
            "Manufacturer P/N": "Manufacturer P/N"

        }

    # -----------------------------------------------------
    # Load Cleaned BOM
    # -----------------------------------------------------

    def load_bom(self):

        logging.info("Loading Cleaned BOM...")

        self.bom_data = load_workbook(
            self.cleaned_file,
            data_only=True
)
        

        sheet = self.bom_data.active

        headers = []

        for cell in sheet[1]:
            headers.append(str(cell.value).strip())

        self.bom_sheet = sheet
        self.headers = headers

        logging.info(
            f"{sheet.max_row - 1} BOM rows loaded."
        )

    # -----------------------------------------------------
    # Open Template
    # -----------------------------------------------------

    def open_template(self):

        logging.info("Opening VSE Template...")

        import os

        print("Template Path:", self.template_file)

        print("File Exists:", os.path.exists(self.template_file))

        self.workbook = load_workbook(self.template_file)

        self.sheet = self.workbook[VSE_SHEET]

    # -----------------------------------------------------
    # Find Header Column
    # -----------------------------------------------------

    def get_column_number(self, header_name):

        header_row = 8

        for cell in self.sheet[header_row]:

            if str(cell.value).strip() == header_name:

                return cell.column

        return None
    
    # -----------------------------------------------------
    # Write BOM Data into VSE Template
    # -----------------------------------------------------

    def write_bom_data(self):

        logging.info("Writing BOM Data to VSE Template...")

        # Create BOM Header Dictionary
        bom_header_index = {}

        for i, header in enumerate(self.headers):
            bom_header_index[header] = i + 1

        current_row = DATA_START_ROW
        template_row = DATA_START_ROW

        # Read each BOM row
        for bom_row in range(2, self.bom_sheet.max_row + 1):

            # Copy template row if more rows are required
            if current_row > self.sheet.max_row:

                self.sheet.insert_rows(current_row)

                for col in range(1, self.sheet.max_column + 1):

                    source = self.sheet.cell(template_row, col)
                    target = self.sheet.cell(current_row, col)

                    if source.has_style:
                        target.font = copy(source.font)
                        target.fill = copy(source.fill)
                        target.border = copy(source.border)
                        target.alignment = copy(source.alignment)
                        target.number_format = copy(source.number_format)
                        target.protection = copy(source.protection)

                    if isinstance(source.value, str) and source.value.startswith("="):
                        target.value = source.value

                self.sheet.row_dimensions[current_row].height = \
                    self.sheet.row_dimensions[template_row].height

            # Write mapped BOM data
            for bom_col, vse_col in self.column_mapping.items():

                if bom_col not in bom_header_index:
                    continue

                source_column = bom_header_index[bom_col]

                value = self.bom_sheet.cell(
                    row=bom_row,
                    column=source_column
                ).value

                target_column = self.get_column_number(vse_col)

                if target_column is not None:
                    self.sheet.cell(
                        row=current_row,
                        column=target_column
                    ).value = value

            current_row += 1

        logging.info("BOM Data Written Successfully.")

    
    # -----------------------------------------------------
    # Save Generated VSE
    # -----------------------------------------------------

    def save(self):

        logging.info("Saving Generated VSE...")

        self.workbook.save(self.generated_vse)

        logging.info("Generated VSE Saved.")

        print(f"\nGenerated VSE File:\n{self.generated_vse}")

        

    # -----------------------------------------------------
    # Run Generator
    # -----------------------------------------------------

    def generate(self):

        logging.info("Starting VSE Generation...")

        self.load_bom()

        self.open_template()

        self.write_bom_data()

        self.save()

        logging.info("VSE Generation Completed.")