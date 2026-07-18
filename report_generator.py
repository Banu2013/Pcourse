"""
=========================================================
BaavaTech BOM Automation Tool
Version : 1.0

Developed by:
BaavaTech

Author:
Banumathi

Copyright © 2026 BaavaTech.
All Rights Reserved.
=========================================================
"""
import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from config import OUTPUT_FOLDER, LOG_FILE

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)


class ReportGenerator:

    def __init__(self):

        self.report_file = os.path.join(
            OUTPUT_FOLDER,
            "Verification_Report.xlsx"
        )

    # --------------------------------------------------
    # Generate Final Review Report
    # --------------------------------------------------

    def generate(self):

        logging.info("Generating Final Review Report...")

    verification_file = os.path.join(
        OUTPUT_FOLDER,
        "Verification_Report.xlsx"
    )

    final_report = os.path.join(
        OUTPUT_FOLDER,
        "Final_Report.xlsx"
    )

    wb = load_workbook(verification_file)

    ws = wb["Verification"]

    green_fill = PatternFill(
        fill_type="solid",
        start_color="92D050"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FF0000"
    )

    bold = Font(bold=True)

    # Header formatting
    for cell in ws[1]:
        cell.font = bold

    # Find Status column
    status_col = None

    for cell in ws[1]:
        if cell.value == "Status":
            status_col = cell.column
            break

    if status_col:
        for row in range(2, ws.max_row + 1):

            value = ws.cell(row, status_col).value

            if value == "PASS":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = green_fill

            elif value == "FAIL":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = red_fill

    wb.save(final_report)

    logging.info("Final Report Generated.")

    print("\nFinal_Report.xlsx created successfully.")